from create_problem import *
import numpy as np
from openpyxl import Workbook, load_workbook
from itertools import repeat


def write_to_spreadsheet(optimal_positions, loadname, savename, match_num, alliance_color):
    print(optimal_positions)

    # Extracts the position values into a list
    position_values_list = [value for (position_name, value) in optimal_positions.items() if
                            'position' in position_name]

    position_values = np.zeros((3, 6))
    # Rearranges the list into a numpy array with with  3 rows representing the scoring location
    # and 6 columns representing the team number/game piece type (two columns per team)
    for (column, index) in zip(range(0, 7), range(3, 19, 3)):
        print(column, index)
        position_values[:, column] = np.column_stack(position_values_list[index - 3:index])
    print(position_values)
    # Rearranges the matrix into 3 columns corresponding to team number and 6 rows:
    # First three rows are scoring locations for cargo (CS, RL, RMH) and last three are for panels
    position_values = np.concatenate((position_values[:, [0, 2, 4]], position_values[:, [1, 3, 5]]), axis=0)
    print(position_values)
    # Finally, rearranges the matrix so that first three rows are for panels rather than cargo
    position_values = np.concatenate((position_values[3:6, :], position_values[0:3, :]), axis=0)
    print(position_values)

    """ 
    =====================================================================================================
                                            Writing to Spreadsheet
    =====================================================================================================
    """
    # Loads the workbook and assigns the worksheet to a variable
    workbook = load_workbook('DeepSpace Scoring Optimizer.xlsx') if loadname == 'default' else load_workbook(loadname)
    sheet = workbook['Match']

    # List of the column letters to to write data in
    columns = ['C', 'D', 'E'] if alliance_color == 'blue' else ['G', 'H', 'I'] if alliance_color == 'red' else None
    # The column the write scoring location totals to
    totals_column = 'F' if alliance_color == 'blue' else 'J' if alliance_color == 'red' else None

    # List of unique teams in the order they're stored in the optimal_positions dictionary
    teams = sorted(set([optimal_position.split('_')[1] for optimal_position in optimal_positions if
                        'position' in optimal_position]))

    # Write the match number to the sheet
    sheet['B1'] = match_num

    # Write each team's number to their corresponding column in the sheet
    for team, column in zip(teams, columns):
        sheet[column + '3'] = int(team)

    # Write the value for each team's scoring position to the \ sheet
    for value, index in zip(np.nditer(position_values), range(0, 18)):
        # Sets the column letter and resets the row number for each different team
        if index % 6 == 0:
            column = columns[int((index / 6))]
            row = 4
        sheet[column + str(row)].value = int(value)
        row += 1

    # Write the optimum score and number of null panels to use to the sheet
    sheet[columns[1] + '11'] = optimal_positions['score']
    sheet[columns[1] + '14'] = optimal_positions['num_null_panels']

    # Write totals for each scoring location
    for row in range(4, 9 + 1):
        sheet[totals_column + str(row)] = '=SUM({}:{})'.format(columns[0] + str(row), columns[2] + str(row))

    workbook.save(savename)
